# ITP Week 3 Day 1 Practice

# import your required modules/methods
import openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# given the following items, using the methods we covered, write to openpyxl

headers = ["id", "name", "base_experience", "height", "order", "weight"]

for col_num, header in enumerate(headers, 1): 
    ws.cell(row=1, column=col_num, value=header)

# use an external counter with just a for loop (no function)
clefairy = {
    "id": 35,
    "name": "clefairy",
    "base_experience": 113,
    "height": 6,
    "order": 56,
    "weight": 75,
}

# method one - no function, looping with range() over column numbers
values_list = list(clefairy.values())
for col_number in range(1,len(values_list)+1):
    ws.cell(
        row = 2,
        column = col_number,
        value = values_list[col_number - 1]
    )

wb.save("week_3\day_1\practiceinventory.xlsx")

# method two - without having to convert to list. "Dictionaries loop over keys by default, so we can use a counter to track columns:"

col_count = 1
for key in clefairy:
    ws.cell(
        row = 3, 
        column = col_count,
        value = clefairy[key]
    )
    col_count += 1

wb.save("week_3\day_1\practiceinventory.xlsx")

# method three - looping with .items - .items() gives both key and value at the same time:

col_count = 1
for key, value in clefairy.items():
    ws.cell(
        row = 4,
        column = col_count, 
        value = value
    )
    col_count += 1

wb.save("week_3\day_1\practiceinventory.xlsx")

# method four - using enumerate() to handle the counter for us

for col_count_en, value in enumerate(clefairy.values(), 1):
    ws.cell(
        row = 5, 
        column = col_count_en,
        value = value
    )

    wb.save("week_3\day_1\practiceinventory.xlsx")

# create a function that takes in a pokemon
weedle = {
    "id": 13,
    "name": "weedle",
    "base_experience": 39,
    "height": 3,
    "order": 17,
    "weight": 32
}

# call the function with weedle!

def addNewPokemon(dict):
    next_row = ws.max_row + 1 # can also use ws.append(listname)
    for col_count_en, value in enumerate(dict.values(), 1):
        ws.cell(
            row = next_row,
            column = col_count_en,
            value = value
        )

addNewPokemon(weedle)

wb.save("week_3\day_1\practiceinventory.xlsx")

# wb.save('./spreadsheets/practice.xlsx')