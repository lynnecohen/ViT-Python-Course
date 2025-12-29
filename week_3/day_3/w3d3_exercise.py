# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation
# we want to make a copy of the Rick and Morty database (which is provided through the api)

# EASY MODE
# import the appropriate modules (you have 3)

import json
import requests
import openpyxl

# set up a workbook and worksheet titled "Rick and Morty Characters"
wb = openpyxl.Workbook()
ws_first = wb.active
ws_first.title = "Rick and Morty Characters"

# # assign a variable 'data' with the returned GET request
character_url = "https://rickandmortyapi.com/api/character"
data = requests.get(character_url)

database = json.loads(data.text)

# Looking over the data
#
# json_formatted = json.dumps(
#     database,
#     indent = 4
# )

# print(json_formatted)

character = database["results"] # the results key contains all the character data

# create the appropriate headers in openpyxl for all of the keys for a single character

first_char = character[0] # first character in the list - dictionary
title_list = [] # set up blank list to hold the title row

for item in first_char: # this syntax returns the keys
    title_list.append(item) #add the header to the list

# print(title_list) Now the title list is complete

for index, item in enumerate(title_list):
    ws_first.cell(
        row = 1,
        column = index + 1,
        value = item
    )


# loop through all of the 'results' of the data to populate the rows and columns for each character

# NOTE: due to the headers, the rows need to be offset by one!

for index, entry in enumerate(character):
    entry_row = index + 2
    for count, item in enumerate(title_list, start=1): # base the loop off the headers, not the dictionary
        # print("\n",item, entry[item])
        entry_value = json.dumps(entry[item]) # accepting storing items as strings because some of them are dicts
        # print("\n\n Entry Row:",entry_row)
        # print("\n\n Entry Column:",count)
        # print("\n\n Value:",entry_value)
        
        ws_first.cell(
            row = entry_row,
            column = count,
            value = entry_value
        )
    

# MEDIUM MODE

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"

# ws_loc = wb.create_sheet("Rick and Morty Locations")
# ws_eps = wb.create_sheet("Rick and Morty Episodes")

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
# "https://rickandmortyapi.com/api/location"
# "https://rickandmortyapi.com/api/episode"

# populate the new worksheets appropriately with all of the data!

# NOTE: don't forget your headers!


# # HARD MODE
# # Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# # Currently, we are only pulling 20 items per api pull!
# # WE WANT EVERYTHING. (contact instructors for office hours if stuck!)


# DEFINING FUNCTIONS TO REUSE

def build_headers_list(firstresult, headers_list):    
	for entry in firstresult: # pull from only first result
		headers_list.append(entry)

def post_headers_to_ws(headers_list, ws_name):
		for index, header in enumerate(headers_list, start = 1):
			ws_name.cell(
				row = 1,
				column = index,
				value = header
			)

def post_entries_to_ws(results, headers_list, ws_name):
	for entry in results:
		next_row = ws_name.max_row + 1
		for index, header in enumerate(headers_list, start = 1):
			ws_name.cell(
				row = next_row,
				column = index,
				value = json.dumps(entry[header])
			)

def get_next_result(api_url, headers_list, ws_name):
	data = requests.get(api_url)
	getdata = data.json()
	info = getdata["info"]
	results = getdata["results"]
	
	post_entries_to_ws(results, headers_list, ws_name)
	print(info)
	if info['next'] is not None:
		get_next_result(info['next'], headers_list, ws_name)

def build_sheet(api_url, ws_title):
    data = requests.get(api_url)
    getdata = data.json()
    info = getdata["info"]
    results = getdata["results"]
    ws_name = wb.create_sheet(ws_title)
	
    headers_list = []
    build_headers_list(results[0], headers_list)
    post_headers_to_ws(headers_list, ws_name)
    post_entries_to_ws(results, headers_list, ws_name)
	
    print(info)
    if info['next'] is not None:
        get_next_result(info['next'], headers_list, ws_name)

build_sheet("https://rickandmortyapi.com/api/episode", "Rick and Morty Episodes")
build_sheet("https://rickandmortyapi.com/api/location", "Rick and Morty Locations")
			


# # NIGHTMARE
# # The inner information for characters, locations, and episodes, references one another through urls
# # ie. for episode 28, it lists all the character but by their url
# # can you use the URLs to make a subsequent call inside your for loops
# # to replace the url with just the appropriate names?
# # NOTE: need to make use of if statements to see if url exists or not
# # (contact instructors for office hours if stuck!)


# # wb.save("./spreadsheets/exercise.xlsx")

wb.save("week_3\spreadsheets\exercise.xlsx")
