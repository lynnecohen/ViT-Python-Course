# ITP Week 2 Day 2 Exercise
import openpyxl

# import the appropriate method from the correct module
from openpyxl import load_workbook

# Import the workbook that you updated in today's practice from
# "./spreadsheets/inventory.xlsx"
wb = load_workbook("week_2\spreadsheets\inventory.xlsx")

print(wb.sheetnames)

# access and store the appropriate worksheet to the variable 'ws'
ws = wb["CURRENT_MONTH_INVENTORY"]

all_rows = list(ws.rows) # Creates a list of rows including formatting
print(all_rows)
print("\n")

all_values = ws.values # Creates a list of values only
print(list(all_values))
print("\n")

# Define a function called add_order_amount that takes in a single parameter called 'row'

def add_order_amount(row):
    max_amount = int(row[2].value)
    reorder_threshold = int(row[3].value)
    quantity = int(row[4].value)

    if(quantity < reorder_threshold):
        order_amount = max_amount - quantity
        ws.cell(row = row[0].row, column = 6, value=order_amount)
        print (f"order amount {order_amount}")    
    # IF the quantity is less or equal to than the threshold,

        # than calculate the order_amount (max_amount - quantity) 
        # assign the value to that row, column 6

# TIP: create variables for quantity, threshold, max_amount that retrieves the values first for cleanliness

# Perform a for..in loop through the range(2, len(inventory.rows))
#   - call the function add_order_amount() passing in the number of the range

for row in all_rows[1:]:
    add_order_amount(row)

wb.save("week_2\spreadsheets\inventory.xlsx")