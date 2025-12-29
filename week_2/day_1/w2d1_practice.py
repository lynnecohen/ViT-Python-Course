# ITP Week 2 Day 1 (In-Class) Practice
import openpyxl

# A1. from the appropriate library, import only the Workbook
from openpyxl import Workbook

# A2. Before anything, we need a workbook to work with...
wb = Workbook()
#print(type(wb))  # <class 'openpyxl.workbook.workbook.Workbook'>
ws_cust = wb.active              # Get the default (active) worksheet
ws_cust.title = "Customer List"    # Rename the default sheet

ws_prod = wb.create_sheet("Product List")         # adds at the end
#ws_sales = wb.create_sheet("Sales", 0)   # adds at the beginning

print("Sheets in the Workbook:",wb.sheetnames)

# A3. We need to interact with a single worksheet.

# A4. assign the value of "First Name" to A1

ws_cust["A1"] = "First Name"
print(ws_cust["A1"].value)

# A5. assign the value of "Last Name" to B1

ws_cust["B1"] = "Last Name"
print(ws_cust["B1"].value)

ws_cust.cell(row=1, column=3, value="Age")
print(ws_cust.cell(row=1, column=3).value)

ws_cust.cell(row=2, column=1, value="Amy")
ws_cust.cell(row=2, column=2, value="Leigh")
ws_cust.cell(row=2, column=3, value="48")

#Doing Math in the sheet
ws_cust["A3"] = 11
ws_cust["B3"] = 4
ws_cust["C3"] = ws_cust["A3"].value + ws_cust["B3"].value

# STOP HERE - RETURN TO LECTURE

# B1. For all of column A, starting at row 2 until row 10, make the cell values: "Gabriel" (attempt a loop)

i = 2
while i < 11: 
    ws_cust.cell(row=i, column=1, value="Gabriel")
    i += 1

# B2. Loop through a range from row 2 to 10 and assign the cell value to last names according to index in column B
# NOTE: PAY ATTENTION to the starting number of the range and how it differs from the starting index of the list

last_names = ['Rolley', 'Smith', 'Balenga', 'Issac', 'Cruise', 'Depp', 'Heard', 'Qiao', 'Biden']

i = 2

for last_name_value in last_names:
    ws_cust.cell(row=i, column=2, value=last_name_value)
    i += 1

# B3. Save the file
# wb.save("./spreadsheets/day_1_practice.xlsx")

wb.save("week_2\spreadsheets\day_1_practice.xlsx")