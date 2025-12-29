import json
import requests
import openpyxl

wb = openpyxl.Workbook()
ws_first = wb.active
ws_first.title = "Rick and Morty Characters"
character_url = "https://rickandmortyapi.com/api/character"
data = requests.get(character_url)

database = json.loads(data.text)

# Looking over the data
#
# json_formatted = json.dumps(
#     database,
#     indent = 4
# )

# print(json_formatted)

character = database["results"] # the results key contains all the character data

first_char = character[0] # first character in the list - dictionary
title_list = [] # set up blank list to hold the title row

for item in first_char: # this syntax returns the keys
    title_list.append(item) #add the header to the list

for index, item in enumerate(title_list):
    ws_first.cell(
        row = 1,
        column = index + 1,
        value = item
    )


# loop through all of the 'results' of the data to populate the rows and columns for each character
for index, entry in enumerate(character):
    entry_row = index + 2
    for count, item in enumerate(title_list, start=1): # base the loop off the headers, not the dictionary
        # print("\n",item, entry[item])
        entry_value = json.dumps(entry[item]) # accepting storing items as strings because some of them are dicts
        # print("\n\n Entry Row:",entry_row)
        # print("\n\n Entry Column:",count)
        # print("\n\n Value:",entry_value)
        
        ws_first.cell(
            row = entry_row,
            column = count,
            value = entry_value
        )
    

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"
# populate the new worksheets appropriately with all of the data!

# DEFINING FUNCTIONS TO REUSE

def build_headers_list(firstresult, headers_list):    
	for entry in firstresult: # pull from only first result
		headers_list.append(entry)

def post_headers_to_ws(headers_list, ws_name):
		for index, header in enumerate(headers_list, start = 1):
			ws_name.cell(
				row = 1,
				column = index,
				value = header
			)

def post_entries_to_ws(results, headers_list, ws_name):
	for entry in results:
		next_row = ws_name.max_row + 1
		for index, header in enumerate(headers_list, start = 1):
			ws_name.cell(
				row = next_row,
				column = index,
				value = json.dumps(entry[header])
			)

def get_next_result(api_url, headers_list, ws_name):
	data = requests.get(api_url)
	getdata = data.json()
	info = getdata["info"]
	results = getdata["results"]
	
	post_entries_to_ws(results, headers_list, ws_name)
	print(info)
	if info['next'] is not None:
		get_next_result(info['next'], headers_list, ws_name)

def build_sheet(api_url, ws_title):
    data = requests.get(api_url)
    getdata = data.json()
    info = getdata["info"]
    results = getdata["results"]
    ws_name = wb.create_sheet(ws_title)
	
    headers_list = []
    build_headers_list(results[0], headers_list)
    post_headers_to_ws(headers_list, ws_name)
    post_entries_to_ws(results, headers_list, ws_name)
	
    print(info)
    if info['next'] is not None:
        get_next_result(info['next'], headers_list, ws_name)

build_sheet("https://rickandmortyapi.com/api/episode", "Rick and Morty Episodes")
build_sheet("https://rickandmortyapi.com/api/location", "Rick and Morty Locations")
			
wb.save("week_3\spreadsheets\exercise.xlsx")
