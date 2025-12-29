import json
import requests
import openpyxl

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "City Weather Report"

url1 = "https://cityweatherclass.free.beeceptor.com/cities" # cities api - provides city, population, region
response1 = requests.get(url1)
print("type",type(response1))
dict_cities = response1.json()
# print("\n\n City Data:",json_cities)

url2 = "https://cityweatherclass.free.beeceptor.com/weather" # city, temperature, humidity
response2 = requests.get(url2)
dict_weather = response2.json()
# print("\n\n Weather Data:",json_weather)

for key, dict in enumerate(dict_cities):
    city_name = dict["city"]

    for cityw in dict_weather:
        weather_city = cityw["city"]
    
        if (weather_city == city_name):
            # print("match!")
            dict["temperature"] = cityw["temperature"]
            dict["humidity"] = cityw["humidity"]

headers = list(dict_cities[0].keys())

for index, header in enumerate(headers):
    ws.cell(
        row = 1,
        column = index + 1,
        value = header.title()
    )

print(dict_cities)

def populateRow(dict_city, row_num):
    for key, val in dict_city.items():
        headertitle = key
        # print(headertitle)
        column_num = headers.index(headertitle) + 1
        # print(column_num)
        # print(val)
        ws.cell(
            row = row_num,
            column = column_num,
            value = val
        )

for key, dict_city in enumerate(dict_cities):
    row_num = key + 2
    print(dict_city, row_num)  
    populateRow(dict_city, row_num)

wb.save("week_4\city_weather_report.xlsx")

# PART 2 STARTS HERE
